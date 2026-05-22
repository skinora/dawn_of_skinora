# Bygger Skinora SEO-søkeordsplan (Fase 1). Kjør: python build_sokeordsplan.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

INK, INK2 = "2D1B3D", "3A3448"
PILAR, BOFU, AMBER, BAND, GRID = "E7E1EF", "F1E5DC", "FFF2CC", "F4F2F7", "D8D5DE"
F = "Arial"
thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ============ ARK 1: Søkeordsplan ============
ws = wb.active
ws.title = "Søkeordsplan"

headers = ["ID", "Type", "Artikkeltittel (foreslått)", "Hovedsøkeord", "Sekundærsøkeord",
           "Søkeintensjon", "Trakt-steg", "Søkevolum (estimat)", "Konkurranse (estimat)",
           "Faktisk søkevolum (fyll inn)", "Internlenking → pilar", "URL-slug (foreslått)",
           "Ordmål", "«Folk spør også» / FAQ-vinkler", "Prioritet", "Status", "Notater / strategi"]

# (id, type, tittel, hoved, sekundaer, intensjon, trakt, volum, konkurranse,
#  pilar, slug, ordmaal, paa, prioritet, notat)
A = [
("P1", "Pilar", "LED-lysterapi for huden: Den komplette guiden",
 "LED lysterapi",
 "LED maske, lysterapi hud, LED ansiktsmaske, rødlysterapi, hva er LED-lysterapi, lysterapimaske",
 "Informasjonell", "TOFU", "Høy", "Høy",
 "— (hub)", "led-lysterapi-guide", "2000–3500",
 "Hva er LED-lysterapi?\nVirker LED-maske egentlig?\nEr LED-maske farlig for øynene?\nHvilke LED-farger finnes og hva gjør de?\nHvor ofte bør man bruke LED-maske?",
 1, "Hovedhub. Publiser først. Skal eie «LED maske»/«LED lysterapi». Lenk ned til alle klynger + til produktsidene Clear og Radiance."),

("P2", "Pilar", "Well-aging: Komplett guide til å eldes med sterk, sunn hud",
 "well-aging",
 "anti-aging vs well-aging, well aging hudpleie, anti-age hudpleie, eldes med huden, hudpleie moden hud",
 "Informasjonell", "TOFU", "Lav (well-aging) / Middels (anti-age)", "Middels",
 "— (hub)", "well-aging-guide", "2000–3500",
 "Hva er well-aging?\nHva er forskjellen på anti-aging og well-aging?\nNår bør man begynne med anti-age?\nHvordan eldes huden?\nKan man bremse aldringstegn?",
 1, "OBS: «well-aging» har lavt søkevolum i Norge – markedet søker «anti-age». Bruk well-aging i tittel/H1 (merkevarevinkling), men dekk «anti-age hudpleie» grundig i H2/brødtekst for å fange volum."),

("P3", "Pilar", "Akne og problemhud: Komplett guide til blått lys og LED-behandling",
 "LED maske akne",
 "blått lys mot kviser, blått lys akne, LED-behandling kviser, lysterapi mot akne, uren hud, LED maske mot kviser",
 "Informasjonell", "TOFU → MOFU", "Middels–Høy", "Høy",
 "— (hub)", "led-maske-akne-blatt-lys", "2000–3500",
 "Virker blått lys mot kviser?\nHjelper LED-maske mot akne?\nHvor lang tid tar det før blått lys virker?\nKan LED-maske gjøre akne verre?\nBlått lys vs. andre aknebehandlinger?",
 1, "Tittel bruker «problemhud» (merkevaretone – se Oversikt, notat 3), men «uren hud»/«kviser» beholdes som søkeord siden det er det folk faktisk søker. Knytt til produktet Skinora Clear."),

("C1", "Klynge", "Rødt lys vs. blått lys: Hva er forskjellen for huden?",
 "rødt lys vs blått lys",
 "forskjell rødt og blått LED-lys, hvilken LED-farge bør jeg velge, rødt lys hud, blått lys hud, LED-farger forklart",
 "Informasjonell", "TOFU", "Middels", "Middels",
 "P1 (primær) + P3", "rodt-lys-vs-blatt-lys", "800–1500",
 "Hva er best av rødt og blått lys?\nKan man kombinere rødt og blått lys?\nHva gjør rødt LED-lys med huden?\nHva gjør blått LED-lys med huden?",
 1, "Grunnleggende – lenk til både P1 og P3. Bør publiseres tidlig som «navet» mellom de to pilarene."),

("C2", "Klynge", "NIR 1072 nm: Hva nær-infrarødt lys er – og hvorfor det er viktig",
 "nær-infrarødt lys",
 "NIR 1072 nm, infrarødt lys hud, NIR lysterapi, bølgelengde LED-maske, 1072 nm",
 "Informasjonell", "TOFU → MOFU", "Veldig lav", "Lav",
 "P1", "nir-1072-nm", "800–1500",
 "Hva er nær-infrarødt lys?\nHva er forskjellen på 1072 nm og 830/850 nm?\nHvor dypt trenger infrarødt lys i huden?\nEr nær-infrarødt lys trygt?",
 2, "Lavt volum, men sterk differensiator – Skinora bruker 1072 nm. Bygger merkevareautoritet (E-E-A-T). Lenk til produkt + P1."),

("C3", "Klynge", "Hvor ofte bør du bruke LED-maske? Frekvens og protokoller",
 "hvor ofte bruke LED-maske",
 "kan man bruke LED-maske hver dag, LED-maske frekvens, hvor lenge LED-maske, LED-maske protokoll",
 "Informasjonell", "MOFU", "Middels", "Middels",
 "P1", "hvor-ofte-bruke-led-maske", "800–1200",
 "Kan man bruke LED-maske hver dag?\nHvor lenge bør en LED-økt vare?\nKan man overdrive LED-behandling?\nNår på dagen bør man bruke LED-maske?",
 1, "Høy MOFU-verdi (folk som eier/vurderer maske). Kilder: 3–5 ganger/uke, 10–20 min, resultater 4–6 uker."),

("C4", "Klynge", "LED-maske med rosacea og sensitiv hud: Trygt eller ikke?",
 "LED-maske rosacea",
 "LED-lys sensitiv hud, rødlysterapi rosacea, LED-maske rødhet, lysterapi sensitiv hud",
 "Informasjonell", "MOFU", "Lav–Middels", "Lav",
 "P3 (primær) + P1", "led-maske-rosacea-sensitiv-hud", "800–1200",
 "Kan man bruke LED-maske med rosacea?\nHjelper rødt lys mot rosacea?\nEr blått lys trygt for sensitiv hud?\nKan LED-maske forverre rødhet?",
 2, "YMYL-nær – vær balansert, anbefal å rådføre hudlege ved diagnose. Eneste dedikerte P3-klynge per nå (se notat 5)."),

("C5", "Klynge", "Kan du bruke LED-maske etter botox og fillere?",
 "LED-maske etter botox",
 "LED-lys etter fillere, rødlysterapi etter botox, hudpleie etter filler, når bruke LED etter botox",
 "Informasjonell", "MOFU", "Lav", "Lav",
 "P2", "led-maske-etter-botox-fillere", "800–1200",
 "Hvor lenge bør man vente med LED etter botox?\nKan rødt lys påvirke fillere?\nEr LED-maske trygt etter ansiktsbehandling?",
 3, "Smal, men kjøpesterk målgruppe (well-aging). Lenk til P2."),

("C6", "Klynge", "LED-maske vs. profesjonell klinikkbehandling: Lønner det seg?",
 "LED-maske vs klinikk",
 "LED-maske hjemme vs klinikk, lønner LED-maske seg, pris LED-behandling klinikk, er LED-maske verdt pengene",
 "Kommersiell", "MOFU → BOFU", "Lav–Middels", "Middels",
 "P1", "led-maske-vs-klinikkbehandling", "1000–1500",
 "Er LED-maske like bra som klinikkbehandling?\nHva koster LED-behandling i klinikk?\nHvor raskt tjener en LED-maske seg inn?",
 2, "Kommersiell vinkel – inkluder ROI-/kostnadsregnestykke og tydelig CTA til produkt."),

("C7", "Klynge", "Hudpleierutine etter 35: Hva endrer seg – og hva du bør gjøre",
 "hudpleie etter 35",
 "hudpleierutine etter 35, hudpleie 40-årene, anti-age rutine, moden hud rutine, hud forandrer seg 35",
 "Informasjonell", "TOFU", "Middels", "Middels",
 "P2", "hudpleierutine-etter-35", "1000–1500",
 "Når begynner huden å eldes?\nHva bør hudpleierutinen inneholde etter 35?\nTrenger jeg retinol etter 35?\nHvor passer LED-maske inn i rutinen?",
 2, "Bredt TOFU-tema – god inngang til P2. Vis hvor LED-maske passer inn i rutinen."),

("C8", "Klynge", "Kombinere LED-lys med retinol, syrer og C-vitamin",
 "LED-maske og retinol",
 "kombinere LED og retinol, LED-lys og C-vitamin, LED og syrer (AHA/BHA), rekkefølge hudpleie LED",
 "Informasjonell", "MOFU", "Lav", "Lav",
 "P2", "kombinere-led-retinol-syrer-c-vitamin", "1000–1500",
 "Kan man bruke retinol og LED-maske sammen?\nBør serum brukes før eller etter LED?\nKan LED-lys bryte ned C-vitamin?\nKan man bruke syrer samme dag som LED?",
 2, "Praktisk «slik gjør du det»-innhold. Lenk til P2 og C3 (frekvens)."),

("C9", "Klynge", "Hvor lang tid tar det å se resultater fra LED-lysterapi?",
 "LED-maske resultater",
 "når ser man effekt av LED, hvor lang tid LED-maske resultater, LED-lys før og etter, virker LED-maske",
 "Informasjonell", "MOFU", "Middels", "Middels",
 "P1", "led-lysterapi-resultater-tid", "800–1200",
 "Når ser man resultater fra LED-maske?\nHvor lenge må man bruke LED-maske før det virker?\nVirker LED-maske med en gang?\nHvorfor ser jeg ikke resultater?",
 1, "Adresserer kjøps-tvil direkte – viktig MOFU. Vær ærlig: kilder oppgir 4–6 uker. Lenk til C14 (vanlige feil)."),

("C10", "Klynge", "Mørke ringer under øynene: Kan LED-lys hjelpe?",
 "mørke ringer under øynene",
 "LED-lys mørke ringer, fjerne mørke ringer, rødt lys øyne, poser under øynene, trøtt hud rundt øynene",
 "Informasjonell", "TOFU", "Middels–Høy", "Middels–Høy",
 "P2", "morke-ringer-under-oynene-led", "800–1200",
 "Hva forårsaker mørke ringer under øynene?\nKan man bli kvitt mørke ringer?\nHjelper rødt lys mot mørke ringer?\nHva er forskjellen på mørke ringer og poser?",
 2, "Bredt søkeord med volum, men temaet tangerer LED – vær ærlig (LED hjelper delvis, ikke mirakelkur)."),

("C11", "Klynge", "Solskader og pigmentering: Kan rødt lys reparere huden?",
 "solskader hud",
 "pigmentering hud, rødt lys pigmentering, LED-lys solskader, pigmentflekker behandling, hyperpigmentering",
 "Informasjonell", "TOFU", "Middels", "Middels",
 "P2", "solskader-pigmentering-rodt-lys", "800–1200",
 "Kan man reparere solskadet hud?\nHjelper rødt lys mot pigmentflekker?\nHva er forskjellen på pigmentflekker og solskader?\nKan LED-maske brukes mot pigment?",
 2, "Knytt til well-aging (fotoaldring). Vær varsom med løfter om «reparere» – nyansér."),

("C12", "Klynge", "LED-maske eller laser: Hva passer for deg?",
 "LED-maske vs laser",
 "LED eller laser hud, forskjell LED og laser, laserbehandling vs LED, hva er best LED eller laser",
 "Kommersiell", "MOFU", "Lav–Middels", "Middels",
 "P1", "led-maske-eller-laser", "1000–1500",
 "Er laser bedre enn LED?\nHva er forskjellen på LED og laser?\nKan LED-maske erstatte laserbehandling?\nHva er tryggest av LED og laser?",
 3, "Sammenligning – posisjonér LED ærlig (skånsom, hjemmebruk) vs laser (kraftigere, klinikk)."),

("C13", "Klynge", "Kollagen, elastin og lys: Vitenskapen forklart enkelt",
 "rødt lys kollagen",
 "kollagen og rødt lys, øke kollagen i huden, elastin hud, stimulere kollagenproduksjon, rødlysterapi kollagen",
 "Informasjonell", "TOFU", "Lav–Middels", "Middels",
 "P2", "kollagen-elastin-rodt-lys", "800–1500",
 "Kan rødt lys øke kollagen?\nHvordan stimulerer man kollagenproduksjon?\nHva er forskjellen på kollagen og elastin?\nNår begynner kollagen å avta?",
 2, "E-E-A-T-innhold: siter fagfellevurderte studier, bruk forfatter med kompetanse. Styrker hele domenets autoritet."),

("C14", "Klynge", "7 vanlige feil folk gjør med LED-masker (og hvordan unngå dem)",
 "vanlige feil LED-maske",
 "LED-maske tips, bruke LED-maske riktig, LED-maske virker ikke, hvorfor virker ikke LED-masken",
 "Informasjonell", "MOFU", "Lav", "Lav",
 "P1", "vanlige-feil-led-maske", "800–1200",
 "Hvorfor virker ikke LED-masken min?\nHvor lenge bør man bruke LED-maske?\nKan man bruke LED-maske med solkrem på?\nMå huden være ren før LED?",
 3, "Listicle – lett delbar, god for internlenking. «Virker ikke»-vinkelen fanger frustrerte eiere."),

("C15", "Klynge", "LED-maske i graviditet og amming: Er det trygt?",
 "LED-maske gravid",
 "LED-lys graviditet, rødlysterapi gravid, LED-maske amming, er LED-maske trygt gravid, hudpleie gravid",
 "Informasjonell", "MOFU", "Lav", "Lav",
 "P1", "led-maske-gravid-amming", "800–1000",
 "Kan man bruke LED-maske når man er gravid?\nEr rødt lys trygt i svangerskapet?\nKan man bruke LED-maske ved amming?\nHva sier ekspertene om LED og graviditet?",
 3, "YMYL-tema – bruk forsiktig språk, ansvarsfraskrivelse, anbefal å rådføre lege/jordmor. Ingen medisinske løfter."),

("B1", "Bunn-trakt", "Skinora Radiance vs. Clear: Hvilken LED-maske passer for deg?",
 "Skinora Radiance vs Clear",
 "Skinora Clear eller Radiance, Skinora sammenligning, hvilken Skinora-maske, Skinora Radiance Face, Skinora Clear",
 "Transaksjonell / Navigasjonell", "BOFU", "Lav (merkevaresøk – vokser med merkevaren)", "Lav (egen merkevare)",
 "P1 + begge produktsider", "skinora-radiance-vs-clear", "1000–1500",
 "Hva er forskjellen på Skinora Radiance og Clear?\nHvilken Skinora-maske er best mot akne?\nHvilken Skinora-maske er best mot aldring?\nKan man bruke begge?",
 1, "Ren BOFU-konvertering. Sammenligningstabell + tydelig CTA til begge produktsider. Lavt volum, men høyest konverteringsverdi."),

("B2", "Bunn-trakt", "Test og sammenligning: De beste LED-maskene i Norge 2026",
 "LED-maske best i test 2026",
 "beste LED-maske Norge, LED-maske test, LED-maske sammenligning, LED ansiktsmaske test, rødlysterapi best i test",
 "Kommersiell", "BOFU", "Høy", "Veldig høy",
 "P1", "led-maske-test-norge-2026", "1500–2500",
 "Hvilken LED-maske er best i test?\nHva bør man se etter i en LED-maske?\nEr dyre LED-masker bedre enn billige?\nHvor mange bølgelengder bør en LED-maske ha?",
 2, "ADVARSEL: «best i test» er svært konkurranseutsatt (mange affiliate-/testsider). Troverdighetsrisiko: egen nettbutikk som rangerer for «test» kan oppfattes som partisk. Vurder en transparent kjøpsguide med åpne kriterier framfor en «test»."),
]

ws["A1"] = "SKINORA — SØKEORDSPLAN  ·  20 artikler  ·  3 pilarer / 15 klyngeartikler / 2 bunn-trakt"
ws.merge_cells("A1:Q1")
ws["A1"].font = Font(name=F, bold=True, size=12, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color=INK)
ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[1].height = 30

for ci, h in enumerate(headers, start=1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=INK2)
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[2].height = 34

for ri, art in enumerate(A, start=3):
    (aid, atype, tittel, hoved, sek, intent, trakt, vol, konk,
     pilar, slug, ordm, paa, prio, notat) = art
    values = [aid, atype, tittel, hoved, sek, intent, trakt, vol, konk,
              "", pilar, slug, ordm, paa, prio, "Ikke startet", notat]
    rowfill = PILAR if atype == "Pilar" else (BOFU if atype == "Bunn-trakt" else None)
    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name=F, size=9, bold=(ci in (1, 2) or atype == "Pilar"))
        c.border = BORDER
        wrap = ci in (3, 5, 11, 14, 17)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if ci in (1, 2, 7, 8, 9, 13, 15) else "left")
        if ci == 10:
            c.fill = PatternFill("solid", start_color=AMBER)
        elif rowfill:
            c.fill = PatternFill("solid", start_color=rowfill)
    ws.row_dimensions[ri].height = 116

widths = [6, 12, 38, 21, 36, 15, 13, 15, 17, 19, 20, 27, 12, 44, 9, 13, 48]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

ws.freeze_panes = "D3"
ws.auto_filter.ref = "A2:Q22"

dv_status = DataValidation(type="list",
    formula1='"Ikke startet,Skisse,Skrives,Til review,Publisert"', allow_blank=False)
dv_prio = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_prio)
dv_status.add("P3:P22")
dv_prio.add("O3:O22")

# ============ ARK 2: Oversikt ============
ov = wb.create_sheet("Oversikt", 0)
ov.sheet_view.showGridLines = False
ov.column_dimensions["A"].width = 3
ov.column_dimensions["B"].width = 46
for col in "CDEFGH":
    ov.column_dimensions[col].width = 14

def section(row, text):
    ov.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c = ov.cell(row=row, column=2, value=text)
    c.font = Font(name=F, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=INK2)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ov.row_dimensions[row].height = 22

def body(row, text, height):
    ov.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c = ov.cell(row=row, column=2, value=text)
    c.font = Font(name=F, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ov.row_dimensions[row].height = height

ov.merge_cells("B2:H2")
ov["B2"] = "SKINORA — SEO-SØKEORDSPLAN (FASE 1)"
ov["B2"].font = Font(name=F, bold=True, size=15, color=INK)
ov.merge_cells("B3:H3")
ov["B3"] = "Strategi og søkeordsplan  ·  pilar–klynge-modell  ·  utarbeidet mai 2026"
ov["B3"].font = Font(name=F, size=10, italic=True, color="6B6577")
ov.row_dimensions[2].height = 24

section(5, "SLIK BRUKER DU ARKET")
body(6, "Dette regnearket er søkeordsplanen for Skinoras innholdssatsing. Arket «Søkeordsplan» "
        "har én rad per artikkel (20 totalt) med hovedsøkeord, sekundærsøkeord og søkeintensjon "
        "— pluss trakt-steg, estimert volum/konkurranse, internlenking, URL-slug, ordmål, "
        "«Folk spør også»-vinkler og status.\n\n"
        "1) Fyll inn den ambergule kolonnen «Faktisk søkevolum» med tall fra Google Keyword "
        "Planner eller Ahrefs.   2) Juster «Prioritet» (1–3) etter volum × intensjon × konkurranse.   "
        "3) Bruk «Status» til å følge produksjonen.   Bruk autofilter i radoverskriften til å "
        "sortere/filtrere.", 118)

section(8, "PILAR–KLYNGE-MODELLEN")
body(9, "Innholdet er bygget som tre PILARSIDER (store, omfattende guider) som hver samler en "
        "gruppe KLYNGEARTIKLER (fokuserte artikler). Hver klynge lenker opp til sin pilar og til "
        "1–2 søsterklynger; hver pilar lenker ned til alle klyngene sine. To BUNN-TRAKT-artikler "
        "fanger kjøpsklare søk. Dette gir Google en tydelig temastruktur og samler "
        "lenkeautoritet på pilarene. Se arket «Internlenking» for kartet.", 88)

section(11, "STATUSOVERSIKT")
dash = [
    ("Artikler totalt", "=COUNTA('Søkeordsplan'!$A$3:$A$22)"),
    ("Pilarsider", "=COUNTIF('Søkeordsplan'!$B$3:$B$22,\"Pilar\")"),
    ("Klyngeartikler", "=COUNTIF('Søkeordsplan'!$B$3:$B$22,\"Klynge\")"),
    ("Bunn-trakt-artikler", "=COUNTIF('Søkeordsplan'!$B$3:$B$22,\"Bunn-trakt\")"),
    ("Prioritet 1 (publiseres først)", "=COUNTIF('Søkeordsplan'!$O$3:$O$22,1)"),
    ("Publisert", "=COUNTIF('Søkeordsplan'!$P$3:$P$22,\"Publisert\")"),
    ("Gjenstår", "=COUNTA('Søkeordsplan'!$A$3:$A$22)-COUNTIF('Søkeordsplan'!$P$3:$P$22,\"Publisert\")"),
]
r = 12
for label, formula in dash:
    lc = ov.cell(row=r, column=2, value=label)
    lc.font = Font(name=F, size=10)
    lc.border = BORDER
    lc.alignment = Alignment(vertical="center", indent=1)
    vc = ov.cell(row=r, column=3, value=formula)
    vc.font = Font(name=F, size=10, bold=True, color=INK)
    vc.fill = PatternFill("solid", start_color=BAND)
    vc.border = BORDER
    vc.alignment = Alignment(vertical="center", horizontal="center")
    ov.row_dimensions[r].height = 19
    r += 1

section(20, "FORKLARING (LEGENDE)")
body(21, "SØKEINTENSJON  —  Informasjonell: vil lære/forstå noe.   Kommersiell: vurderer og "
         "sammenligner alternativer.   Transaksjonell: klar til å kjøpe.   Navigasjonell: søker "
         "en bestemt merkevare/side.\n\n"
         "TRAKT-STEG  —  TOFU: toppen av trakten (oppmerksomhet/lær).   MOFU: midten "
         "(vurdering).   BOFU: bunnen (kjøpsbeslutning).\n\n"
         "VOLUM & KONKURRANSE  —  kvalitativ skala: Veldig lav · Lav · Middels · Høy · Veldig "
         "høy. Dette er ESTIMATER basert på markeds- og SERP-kunnskap, ikke måledata "
         "(se notat 1).\n\n"
         "FARGER  —  Lilla rad = pilarside.   Beige rad = bunn-trakt-artikkel.   "
         "Amber kolonne = felt du skal fylle inn.", 168)

section(23, "VIKTIGE STRATEGISKE NOTATER")
body(24,
 "1.  SØKEVOLUM MANGLER.  Kolonne H/I er kvalitative estimater — IKKE live data fra Ahrefs/"
 "Keyword Planner (jeg har ikke tilgang til de verktøyene). Fyll inn faktiske tall i den "
 "ambergule kolonnen «Faktisk søkevolum» før endelig prioritering.\n\n"
 "2.  «WELL-AGING» (P2) har lavt søkevolum i Norge — markedet søker «anti-age». Behold "
 "well-aging som merkevarevinkling i tittel/H1, men dekk «anti-age hudpleie» grundig i "
 "brødteksten for å fange volumet.\n\n"
 "3.  «UREN HUD» vs «PROBLEMHUD» (P3).  Skinoras tone unngår dømmende ord som «uren hud» "
 "i kundevendt tekst. «Uren hud» beholdes likevel som SØKEORD fordi det er det folk faktisk "
 "søker. Praksis: bruk «problemhud» i synlige overskrifter; la «uren hud»/«kviser» dukke opp "
 "naturlig i brødtekst, meta-tagger og FAQ der søkeordet trengs.\n\n"
 "4.  «BEST I TEST» (B2) er svært konkurranseutsatt og en troverdighetsrisiko: en nettbutikk "
 "som rangerer for «test» av sin egen produktkategori kan oppfattes som partisk. Vurder en "
 "transparent kjøpsguide med åpne kriterier framfor en «test».\n\n"
 "5.  P3 (AKNE) ER TYNN på dedikerte klynger — kun C4 er primært knyttet hit. Vurder 2–3 "
 "ekstra akne-klynger i neste pulje, f.eks. «LED mot aknearr», «blått lys mot hormonell akne», "
 "«LED-maske mot store porer».\n\n"
 "6.  INTERNLENKING.  Hver klynge lenker opp til sin pilar og til 1–2 søsterklynger; hver "
 "pilar lenker ned til alle klyngene sine. Lenk til produktsidene (Clear/Radiance) der det er "
 "naturlig. Se arket «Internlenking».\n\n"
 "7.  E-E-A-T.  Oppgi forfatter med relevant kompetanse, siter fagfellevurderte studier "
 "(særlig P1, P3 og C13), og hold publiseringsdato oppdatert. YMYL-temaer (C4, C15) krever "
 "ekstra forsiktig språk og ansvarsfraskrivelse.", 360)

section(26, "NESTE STEG")
body(27, "1.  Fyll inn «Faktisk søkevolum» fra Google Keyword Planner / Ahrefs.\n"
         "2.  Juster «Prioritet» etter volum × søkeintensjon × konkurranse.\n"
         "3.  Start produksjon med Prioritet 1: de tre pilarene + C1, C3, C9 og B1.\n"
         "4.  Publiser pilarene først, deretter klyngene som lenker til dem.\n"
         "5.  Følg produksjonen i «Status»-kolonnen; gå videre til Fase 2 (innholdsproduksjon).", 104)

# ============ ARK 3: Internlenking ============
il = wb.create_sheet("Internlenking")
il.sheet_view.showGridLines = False
il_widths = [3, 34, 34, 20, 30, 34]
for i, w in enumerate(il_widths, start=1):
    il.column_dimensions[chr(64 + i)].width = w

il.merge_cells("B2:F2")
il["B2"] = "INTERNLENKING — PILAR/KLYNGE-KART"
il["B2"].font = Font(name=F, bold=True, size=13, color=INK)
il.row_dimensions[2].height = 22

il_head = ["Pilar (hub)", "Klynger som lenker primært hit", "Også relevant for", "Produktside(r)", "Merknad"]
for ci, h in enumerate(il_head, start=2):
    c = il.cell(row=4, column=ci, value=h)
    c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=INK2)
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    c.border = BORDER
il.row_dimensions[4].height = 30

il_rows = [
    ("P1 — LED-lysterapi for huden", "C1, C2, C3, C6, C9, C12, C14, C15", "B1, B2",
     "Skinora Clear + Radiance Face", "Generell hub – binder hele klyngen sammen."),
    ("P2 — Well-aging", "C5, C7, C8, C10, C11, C13", "C2",
     "Skinora Radiance Face", "Aldring / moden hud."),
    ("P3 — Akne og problemhud", "C4", "C1, C3, C9",
     "Skinora Clear", "Tynn på dedikerte klynger – se notat 5 i Oversikt."),
]
for ri, row in enumerate(il_rows, start=5):
    for ci, val in enumerate(row, start=2):
        c = il.cell(row=ri, column=ci, value=val)
        c.font = Font(name=F, size=9, bold=(ci == 2))
        c.fill = PatternFill("solid", start_color=PILAR if ci == 2 else "FFFFFF")
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    il.row_dimensions[ri].height = 40

il.merge_cells("B9:F9")
il["B9"] = "LENKEPRINSIPPER"
il["B9"].font = Font(name=F, bold=True, size=11, color="FFFFFF")
il["B9"].fill = PatternFill("solid", start_color=INK2)
il["B9"].alignment = Alignment(vertical="center", indent=1)
il.row_dimensions[9].height = 22
il.merge_cells("B10:F10")
il["B10"] = ("•  Hver klyngeartikkel lenker OPP til sin pilar med beskrivende ankertekst "
             "(ikke «klikk her»).\n"
             "•  Hver pilar lenker NED til alle klyngene sine — gjerne fra en innholdsfortegnelse "
             "eller «les mer»-seksjoner.\n"
             "•  Klynger lenker til 1–2 relevante søsterklynger (f.eks. C3 ↔ C9 ↔ C14).\n"
             "•  Lenk til produktsidene (Clear/Radiance) der det er naturlig — særlig fra "
             "kommersielle/MOFU-artikler og begge bunn-trakt-artiklene.\n"
             "•  Bunn-trakt-artiklene (B1, B2) skal ha tydelige CTA-er til produktsidene.")
il["B10"].font = Font(name=F, size=10)
il["B10"].alignment = Alignment(vertical="top", wrap_text=True)
il.row_dimensions[10].height = 130

wb.calculation.fullCalcOnLoad = True
wb.save(r"C:\Users\am\dev\dawn_of_skinora\seo\Skinora-SEO-sokeordsplan-Fase1.xlsx")
